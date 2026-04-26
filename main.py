import pandas as pd
import matplotlib.pyplot as plt

from model import train_model, evaluate_model
from utils import preprocess

# Load data
train = pd.read_csv("data/train.csv")
test = pd.read_csv("data/test.csv")

# Preprocess
X_train = preprocess(train)
y_train = train['SalePrice']

X_test = preprocess(test)

# Train model
model = train_model(X_train, y_train)

# Predictions
predictions = model.predict(X_test)

# =========================
# 📄 1. Submission Data
# =========================
submission = pd.DataFrame({
    "Id": test["Id"],
    "SalePrice": predictions
})

# =========================
#2. Future Prices (+10%)
# =========================
future_prices = predictions * 1.10

future_df = pd.DataFrame({
    "Id": test["Id"],
    "FuturePrice": future_prices
})

# =========================
#3. Best Houses Under Budget
# =========================
budget = 200000

best_houses = train[train['SalePrice'] <= budget]
best_houses = best_houses[['Id', 'GrLivArea', 'SalePrice']].head(10)

# =========================
#4. Model Metrics
# =========================
score = evaluate_model(model, X_train, y_train)

metrics_df = pd.DataFrame({
    "Metric": ["R2 Score"],
    "Value": [score]
})

# =========================
#5. Create Graphs
# =========================

# Scatter Plot
plt.scatter(train['GrLivArea'], train['SalePrice'])
plt.xlabel("Living Area")
plt.ylabel("Sale Price")
plt.title("Area vs Price")
plt.savefig("area_price.png")
plt.close()

# Feature Importance
importance = model.feature_importances_
features = X_train.columns

plt.barh(features, importance)
plt.xlabel("Importance")
plt.title("Feature Importance")
plt.savefig("feature_importance.png")
plt.close()

# =========================
#6. Save to Excel (MULTIPLE SHEETS)
# =========================
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Step 1: Create Excel file
with pd.ExcelWriter("output.xlsx") as writer:
    submission.to_excel(writer, sheet_name="Predictions", index=False)
    future_df.to_excel(writer, sheet_name="Future Prices", index=False)
    best_houses.to_excel(writer, sheet_name="Best Houses", index=False)
    metrics_df.to_excel(writer, sheet_name="Model Metrics", index=False)

# Step 2: Add graphs into Excel
wb = load_workbook("output.xlsx")

# Create new sheet
ws = wb.create_sheet("Graphs")

# Insert images
img1 = Image("area_price.png")
img2 = Image("feature_importance.png")

ws.add_image(img1, "A1")
ws.add_image(img2, "A20")

# Save final file
wb.save("output.xlsx")

print("output.xlsx created with predictions, insights, and graphs!")